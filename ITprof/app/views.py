from django.shortcuts import render
from django.http import HttpResponse
import app.models
from pathlib import Path
import openpyxl
import os

#  1. Главная
#  2. Востребованность
#  3. География
#  4. Навыки
#  5. Последние вакансии

BASE_DIR = str(Path(__file__).resolve().parent.parent).replace('\\', '/') + '/app/'

def index(request):
    print(BASE_DIR)
    data = app.models.HomePageInfo.objects.all()[0]
    image = str(data.image)[4:]
    print(image)
    return render(request, 'index.html', {
        'title':data.title,
        'description':data.description,
        'image':image
    })
 
def relevance(request):
   
    data = app.models.RelevancePageInfo.objects.all()[0]
    
    # SALARY
    SalaryName = os.path.basename(f'{data.SalaryByYears_Table}')
    SalaryName = os.path.splitext(SalaryName)[0].replace('_', ' ')

    SalaryGraph = str(data.SalaryByYears_Graph)
    SalaryGraph = SalaryGraph[4:]

    
    SalaryTableFile = BASE_DIR + str(data.SalaryByYears_Table)[4:].replace('\\', '/')
    wookbook = openpyxl.load_workbook(SalaryTableFile)
    worksheet = wookbook.active
    SalaryTable = []
    for i in range(0, worksheet.max_row):
        SalaryTable.append( list( col[i].value for col in worksheet.iter_cols(1, worksheet.max_column) ) )
 
    #COUNT VAC
    CountName = os.path.basename(f'{data.CountByYears_Table}')
    CountName = os.path.splitext(CountName)[0].replace('_', ' ')

    CountGraph = str(data.CountByYears_Graph)
    CountGraph = CountGraph[4:]

    CountTableFile = BASE_DIR + str(data.CountByYears_Table)[4:].replace('\\', '/')
    wookbook = openpyxl.load_workbook(CountTableFile)
    worksheet = wookbook.active
    CountTable = []
    for i in range(0, worksheet.max_row):
        CountTable.append( list( col[i].value for col in worksheet.iter_cols(1, worksheet.max_column) ) )

    return render(request, 'relevance.html', {
        'SalaryName': SalaryName,
        'SalaryGraph': SalaryGraph,
        'SalaryTable': SalaryTable,
        'CountName': CountName,
        'CountGraph': CountGraph,
        'CountTable': CountTable
    })
 
def geography(request):
    data = app.models.GeographyPageInfo.objects.all()[0]
    
    # SALARY
    SalaryName = os.path.basename(f'{data.SalaryByCities_Table}')
    SalaryName = os.path.splitext(SalaryName)[0].replace('_', ' ')

    SalaryGraph = str(data.SalaryByCities_Graph)
    SalaryGraph = SalaryGraph[4:]

    
    SalaryTableFile = BASE_DIR + str(data.SalaryByCities_Table)[4:].replace('\\', '/')
    wookbook = openpyxl.load_workbook(SalaryTableFile)
    worksheet = wookbook.active
    SalaryTable = []
    for i in range(0, worksheet.max_row):
        SalaryTable.append( list( col[i].value for col in worksheet.iter_cols(1, worksheet.max_column) ) )
 
    #COUNT VAC
    CountName = os.path.basename(f'{data.CountByCities_Table}')
    CountName = os.path.splitext(CountName)[0].replace('_', ' ')

    CountGraph = str(data.CountByCities_Graph)
    CountGraph = CountGraph[4:]

    CountTableFile = BASE_DIR + str(data.CountByCities_Table)[4:].replace('\\', '/')
    wookbook = openpyxl.load_workbook(CountTableFile)
    worksheet = wookbook.active
    CountTable = []
    for i in range(0, worksheet.max_row):
        CountTable.append( list( col[i].value for col in worksheet.iter_cols(1, worksheet.max_column) ) )

    return render(request, 'geography.html', {
        'SalaryName': SalaryName,
        'SalaryGraph': SalaryGraph,
        'SalaryTable': SalaryTable,
        'CountName': CountName,
        'CountGraph': CountGraph,
        'CountTable': CountTable
    })

def skills(request):
    data = app.models.SkillsPageInfo.objects.all()[0]

    SkillsName = os.path.basename(f'{data.Skills_Table}')
    SkillsName = os.path.splitext(SkillsName)[0].replace('_', ' ')

    Skills_TableFile = BASE_DIR + str(data.Skills_Table)[4:].replace('\\', '/')
    wookbook = openpyxl.load_workbook(Skills_TableFile)
    worksheet = wookbook.active
    Skills_Table = []
    for i in range(0, worksheet.max_row):
        Skills_Table.append( list( col[i].value for col in worksheet.iter_cols(1, worksheet.max_column) ) )

    return render(request, 'skills.html', {
        'SkillsTable': Skills_Table
    })
    
def last_vacansies(request):
    return render(request, 'last_vacansies.html')

def error(request):
    return render(request, 'error.html')